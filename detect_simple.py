import tensorflow as tf
import core.utils as utils
from tensorflow.python.saved_model import tag_constants
import cv2
import numpy as np
import openpyxl 
from datetime import date
from PIL import Image
from openpyxl.cell import cell
import shutil
import os #모듈 import 

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3' #enable 오류
MODEL_PATH = './checkpoints/yolov4-416' # 모델 기본 위치
IOU_THRESHOLD = 0.45  # 한계점 설정
SCORE_THRESHOLD = 0.40 # 최소 점수 만족
INPUT_SIZE = 416 #사진 크기

# load model
saved_model_loaded = tf.saved_model.load(MODEL_PATH, tags=[tag_constants.SERVING]) # 학습된 모델 부르기
infer = saved_model_loaded.signatures['serving_default'] #시그니쳐 키 설정

tday = date.today()
tday_w = tday.strftime('%Y-%b-%d-%A')
Exc_Location = './excel/result_excel'+"("+tday_w+")"+'.xlsx' #엑셀 링크

def img_detect(img_path):
    
    img = cv2.imread(img_path) # img road
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB) # bgr 형식 rgb로 convert

    img_input = cv2.resize(img, (INPUT_SIZE, INPUT_SIZE)) # 크기 재설정
    img_input = img_input / 255. # 크기 재설정
    img_input = img_input[np.newaxis, ...].astype(np.float32) # 축 설정
    img_input = tf.constant(img_input) # numpy를 tf 로 변환

    pred_bbox = infer(img_input) #  모델에 img 추가

    wb = openpyxl.Workbook()
    wb.active.title = "bounding_boxes" #excel file 설정
    wsheet = wb["bounding_boxes"] 
    wsheet.cell(1,1).value = "classes"
    wsheet.cell(1,2).value = "score"
    wsheet.cell(1,3).value = "UL_x"
    wsheet.cell(1,4).value = "UL_y"
    wsheet.cell(1,5).value = "DR_x"
    wsheet.cell(1,6).value = "DR_y"


    for key, value in pred_bbox.items():
        boxes = value[:, :, 0:4]
        pred_conf = value[:, :, 4:]

    boxes, scores, classes, valid_detections = tf.image.combined_non_max_suppression( #nms 진행 
        boxes=tf.reshape(boxes, (tf.shape(boxes)[0], -1, 1, 4)),
        scores=tf.reshape(
            pred_conf, (tf.shape(pred_conf)[0], -1, tf.shape(pred_conf)[-1])), #모델 정확도 측정
        max_output_size_per_class=50,
        max_total_size=50,
        iou_threshold=IOU_THRESHOLD,
        score_threshold=SCORE_THRESHOLD
    )
    #iou, score 넘는 애들만 체크 

    pred_bbox = [boxes.numpy(), scores.numpy(), classes.numpy(), valid_detections.numpy()]
    result,result_exc = utils.draw_bbox(img, pred_bbox) #바운딩 박스 추가
    for i in range(len(result_exc)):
        for j in range(len(result_exc[i])):
            wsheet.cell(i+2,j+1).value = result_exc[i][j] #엑셀 설정
    wsheet.cell(1,8).value = len(result_exc)
    

    result = cv2.cvtColor(np.array(result), cv2.COLOR_RGB2BGR) # 재변환
    cv2.imwrite('res.png', result) # 사진 저장
    wb.save(Exc_Location)


def img_cut(Exc_Location):
    img = cv2.imread('./data/central.jpg')
    shutil.rmtree('./cut_img') #자른 사진 저장 폴더 초기화
    os.makedirs('./cut_img')

    wb = openpyxl.load_workbook(Exc_Location)
    ws = wb["bounding_boxes"]
    object_cnt = ws.cell(1,8).value # (1,8)에는 항상 감지한 물체들의 개수가 써져있음


    for i in range(2,object_cnt+2):
        cropped_img = img[ws.cell(i,3).value:ws.cell(i,4).value,
            ws.cell(i,5).value:ws.cell(i,6).value].copy() #사진 자르기
        cv2.imwrite('./cut_img/'+str(i-1)+
            '.jpg',cropped_img) #사진 저장
    wb.close()


if __name__ == '__main__':
    img_path = './data/central.jpg'
    img_detect(img_path)
    img_cut(Exc_Location)