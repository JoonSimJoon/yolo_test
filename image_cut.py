from PIL import Image
import openpyxl
from openpyxl.cell import cell
import cv2
from datetime import date
import openpyxl 
import shutil
import os

def main(Exc_Location):
    img = cv2.imread('./data/central.jpg')
    shutil.rmtree('./cut_img') #자른 사진 저장 폴더 초기화
    os.makedirs('./cut_img')

    wb = openpyxl.load_workbook(Exc_Location)
    ws = wb["bounding_boxes"]
    object_cnt = ws.cell(1,8).value # (1,8)에는 항상 감지한 물체들의 개수가 써져있음


    for i in range(2,object_cnt+2):
        cropped_img = img[ws.cell(i,3).value:ws.cell(i,4).value,
            ws.cell(i,5).value:ws.cell(i,6).value].copy() #사진 자르기
        cv2.imwrite('./cut_img/'+str(i-1)+"_"+cell(i,1).value+
            '.jpg',cropped_img) #사진 저장
    wb.close()


if __name__ == '__main__':    
    tday = date.today()
    tday_w = tday.strftime('%Y-%b-%d-%A')
    Exc_Location = './excel/result_excel'+"("+tday_w+")"+'.xlsx' #엑셀 링크
    main(Exc_Location)